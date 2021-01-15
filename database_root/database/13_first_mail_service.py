
import pymysql
import FinanceDataReader as fdr
from datetime import datetime
# from pandas_datareader import wb
import pandas_datareader.data as web
import pandas_datareader as pdr
import datetime
import requests
import json
import pandas as pd
import smtplib
from account import *
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.axis import DateAxis


def macro_value(code):
    df = pdr.DataReader(code, 'fred', start='1990-01-01')
    # df.to_excel('inflation.xlsx')
    return df

def get_macro_excel():
    df1 = macro_value('DGS10')  # 미국 10년 국채 금리
    df2 = macro_value('T10Y2Y') # 미국 10년 국채 금리 - 미국2년 국채 금리(장단기 금리차)
    df3 = macro_value('T10YIE') # 미국 10년 국채 금리 - 미국 물가연동채 = 기대 인플레이션
    df = pd.concat([df1, df2, df3], axis=1)
    # df["date"] = df.indexd
    df.reset_index(inplace=True)
    df.columns = ["date", "미국_10년_국채금리", "10년-2년(장단기금리차)", "기대_인플레이션"]
    df = df[::-1] # 상하 반전
    se1 = df.iloc[0]
    df.to_excel(r"C:\Users\yungu\Desktop\macro_excel\macro_{}.xlsx".format(datetime.datetime.today().strftime("%Y_%m_%d")))
    return se1

def content_fill():
    se1 = get_macro_excel()
    content = "{}===========>미국10년국채금리 : {}, 장단기금리차 : {}, 기대 인플레이션 : {}".format(se1[0], se1[1], se1[2], se1[3])
    return content

def content_fill_filtered():
    se1 = get_macro_excel()
    content = "{}===========>미국10년국채금리 : {}, 장단기금리차 : {}, 기대 인플레이션 : {}".format(se1[0], se1[1], se1[2], se1[3])
    if float(se1[1]) >= 1.5 or float(se1[3]) >= 2.5:
        send_message(content)
    else:
        pass



def send_message(con):
    content = con
    msg = EmailMessage()
    msg["Subject"] = "매크로 지표"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = "Theonpharm@naver.com, juho0124@naver.com, yungun3@naver.com"
    msg.set_content(content)

    with open((r"C:\Users\yungu\Desktop\macro_excel\macro_{}.xlsx".format(datetime.datetime.today().strftime("%Y_%m_%d"))), "rb") as f:
        msg.add_attachment(f.read(), maintype="applocation", subtype="octect-stream", filename=f.name)

    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.ehlo()
        smtp.starttls()
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)
# 그래프 포기...

# def get_macro_graph_to_excel():
#     wb = load_workbook(r"C:\Users\yungu\Desktop\macro_excel\macro_{}.xlsx".format(datetime.datetime.today().strftime("%Y_%m_%d")))
#     ws = wb.active
#     line_value = Reference(ws, min_row=1, max_row=ws.max_row+1, min_col=2, max_col=5)
#     line_chart = LineChart()
#
#     line_chart.add_data(line_value, titles_from_data=True)
#     line_chart.set_categories(line_value)
#     line_chart.title = "거시지표"
#     line_chart.style = 12
#     line_chart.x_axis = DateAxis(crossAx=100)
#     line_chart.x_axis.number_format = 'yy-mmm'
#     line_chart.x_axis.majorTimeUnit = "months"
#     line_chart.y_axis.title= "금리수준"
#     line_chart.x_axis.title= "날짜"
#     ws.add_chart(line_chart, "G11")
#
#     wb.save(r"C:\Users\yungu\Desktop\macro_excel\macro_{}_1.xlsx".format(datetime.datetime.today().strftime("%Y_%m_%d")))






if __name__ == '__main__':
    content_fill_filtered()
    # get_macro_excel()
    # get_macro_graph_to_excel()